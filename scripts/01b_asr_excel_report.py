#!/usr/bin/env python3
"""
Script 01b – ASR Excel Report
==============================
Reads the existing ASR result CSVs (already produced by the evaluation runs)
and builds a comprehensive, multi-sheet Excel workbook + publication-quality
figures.  Does NOT re-run any ASR model.

Usage
-----
  python scripts/01b_asr_excel_report.py

Output
------
  data/outputs/asr_report/asr_full_report.xlsx
  data/outputs/asr_report/figures/
"""

from __future__ import annotations

import logging
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")
logging.getLogger("absl").setLevel(logging.WARNING)

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("asr_report")

OUT_DIR  = ROOT / "data" / "outputs" / "asr_report"
FIG_DIR  = OUT_DIR / "figures"
EXCEL    = OUT_DIR / "asr_full_report.xlsx"

ASR_DIR  = ROOT / "data" / "outputs" / "full_dataset_analysis" / "test"


# ─────────────────────────────────────────────────────────────────────────────
# Load
# ─────────────────────────────────────────────────────────────────────────────

def load_all() -> pd.DataFrame:
    frames = []

    # Whisper models
    for model in ("whisper_tiny", "whisper_base", "whisper_small"):
        p = ASR_DIR / f"{model}_results.csv"
        if not p.exists():
            logger.warning("Missing: %s", p); continue
        df = pd.read_csv(p)
        df = df.rename(columns={"model_name": "model"})
        df["model"] = model
        frames.append(df)
        logger.info("Loaded %d rows – %s", len(df), model)

    # Wav2Vec2
    p = ASR_DIR / "wav2vec2_german" / "test_wav2vec2_german_results.csv"
    if p.exists():
        df = pd.read_csv(p)
        df["model"] = "wav2vec2_german"
        frames.append(df)
        logger.info("Loaded %d rows – wav2vec2_german", len(df))
    else:
        logger.warning("Missing wav2vec2 results")

    if not frames:
        raise FileNotFoundError("No ASR CSVs found. Run the ASR evaluation first.")

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.dropna(subset=["reference_text", "asr_text"])
    combined = combined[combined["asr_text"].str.strip().str.len() > 0]
    logger.info("Total rows: %d", len(combined))
    return combined


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _autofit(ws):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)


def _write_df(ws, df, title=None, freeze="A2"):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    H_FILL = PatternFill("solid", fgColor="1F4E79")
    H_FONT = Font(color="FFFFFF", bold=True, size=10)
    T_FONT = Font(bold=True, size=13)

    if title:
        ws.cell(1, 1, title).font = T_FONT
        start = 3
    else:
        start = 1

    for ri, row in enumerate(dataframe_to_rows(df, index=False, header=True), start):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            if ri == start:
                cell.font  = H_FONT
                cell.fill  = H_FILL
                cell.alignment = Alignment(horizontal="center")
    if freeze:
        ws.freeze_panes = freeze
    _autofit(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Compute extra metrics if not already present
# ─────────────────────────────────────────────────────────────────────────────

def _add_cer(df: pd.DataFrame) -> pd.DataFrame:
    """Character Error Rate — jiwer or simple fallback."""
    if "cer" in df.columns:
        return df
    try:
        import jiwer
        cers = []
        for _, row in df.iterrows():
            ref = str(row.get("reference_text", "")).strip()
            hyp = str(row.get("asr_text", "")).strip()
            if ref and hyp:
                try:
                    cers.append(jiwer.cer(ref, hyp))
                except Exception:
                    cers.append(None)
            else:
                cers.append(None)
        df["cer"] = cers
    except ImportError:
        pass
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(df: pd.DataFrame):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    agg_cols = {}
    for col in ("wer", "medical_wer", "confidence", "processing_time", "audio_duration"):
        if col in df.columns:
            agg_cols[f"{col}_mean"] = (col, "mean")
            agg_cols[f"{col}_std"]  = (col, "std")
            agg_cols[f"{col}_min"]  = (col, "min")
            agg_cols[f"{col}_max"]  = (col, "max")
    if "cer" in df.columns:
        agg_cols["cer_mean"] = ("cer", "mean")
        agg_cols["cer_std"]  = ("cer", "std")

    summary = df.groupby("model").agg(**agg_cols).reset_index()

    # RTF = processing_time / audio_duration
    if "processing_time_mean" in summary.columns and "audio_duration_mean" in summary.columns:
        summary["rtf_mean"] = (summary["processing_time_mean"] / summary["audio_duration_mean"]).round(4)

    summary = summary.sort_values("wer_mean")
    summary.insert(0, "rank", range(1, len(summary) + 1))
    float_cols = [c for c in summary.columns if summary[c].dtype == float]
    summary[float_cols] = summary[float_cols].round(4)
    _write_df(ws, summary, title="ASR Evaluation – Summary (ranked by WER)", freeze="A4")

    # ── Sheet 2: Per-Sample ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Sample")
    cols = ["model", "sample_id", "wer", "cer", "medical_wer", "confidence",
            "processing_time", "audio_duration", "word_count"]
    cols = [c for c in cols if c in df.columns]
    per = df[cols].sort_values(["model", "sample_id"])
    float_c = [c for c in per.columns if per[c].dtype == float]
    per[float_c] = per[float_c].round(4)
    _write_df(ws2, per, freeze="A2")

    # ── Sheet 3: WER Distribution Bins ───────────────────────────────────────
    ws3 = wb.create_sheet("WER-Bins")
    bins  = [0, 0.1, 0.2, 0.3, 0.5, 1.0, float("inf")]
    labels = ["0–10%", "10–20%", "20–30%", "30–50%", "50–100%", ">100%"]
    df["wer_bin"] = pd.cut(df["wer"], bins=bins, labels=labels, right=True)
    bin_df = (
        df.groupby(["model", "wer_bin"], observed=True)
          .size().rename("count")
          .reset_index()
    )
    bin_pivot = bin_df.pivot(index="model", columns="wer_bin", values="count").fillna(0).astype(int)
    bin_pivot["total"] = bin_pivot.sum(axis=1)
    for lbl in labels:
        if lbl in bin_pivot.columns:
            bin_pivot[f"{lbl} %"] = (bin_pivot[lbl] / bin_pivot["total"] * 100).round(1)
    bin_pivot = bin_pivot.reset_index()
    _write_df(ws3, bin_pivot, title="WER Distribution by Bucket", freeze="A2")

    # ── Sheet 4: Audio Duration Analysis ─────────────────────────────────────
    ws4 = wb.create_sheet("Duration-Analysis")
    if "audio_duration" in df.columns:
        dur_bins   = [0, 5, 10, 20, 40, float("inf")]
        dur_labels = ["0-5s", "5-10s", "10-20s", "20-40s", ">40s"]
        df["dur_bin"] = pd.cut(df["audio_duration"], bins=dur_bins, labels=dur_labels, right=True)
        dur_df = (
            df.groupby(["model", "dur_bin"], observed=True)
              .agg(mean_wer=("wer", "mean"), count=("wer", "count"))
              .reset_index()
        )
        dur_df["mean_wer"] = dur_df["mean_wer"].round(4)
        _write_df(ws4, dur_df, title="Mean WER by Audio Duration", freeze="A2")

    # ── Sheet 5: Statistical Comparison ──────────────────────────────────────
    ws5 = wb.create_sheet("Statistical")
    models = df["model"].unique().tolist()
    stat_rows = []
    try:
        from scipy import stats
        for i, m1 in enumerate(models):
            for m2 in models[i+1:]:
                w1 = df[df["model"] == m1]["wer"].dropna().values
                w2 = df[df["model"] == m2]["wer"].dropna().values
                n  = min(len(w1), len(w2))
                if n < 10:
                    continue
                stat, pval = stats.wilcoxon(w1[:n], w2[:n])
                stat_rows.append({
                    "model_A": m1, "model_B": m2,
                    "wer_mean_A": round(w1.mean(), 4),
                    "wer_mean_B": round(w2.mean(), 4),
                    "delta_wer":  round(w1.mean() - w2.mean(), 4),
                    "wilcoxon_stat": round(float(stat), 2),
                    "p_value":  round(float(pval), 6),
                    "significant": "YES" if pval < 0.05 else "NO",
                })
    except ImportError:
        stat_rows = [{"note": "scipy not installed – install with: pip install scipy"}]
    stat_df = pd.DataFrame(stat_rows)
    _write_df(ws5, stat_df, title="Pairwise Wilcoxon Signed-Rank Tests (WER)", freeze="A4")

    # ── Sheet 6: Transcription Samples ───────────────────────────────────────
    ws6 = wb.create_sheet("Sample-Transcriptions")
    rng  = np.random.default_rng(42)
    ids  = rng.choice(df["sample_id"].unique(), size=min(30, df["sample_id"].nunique()), replace=False)
    samp = df[df["sample_id"].isin(ids)].sort_values(["sample_id", "model"])
    samp_cols = ["sample_id", "model", "wer", "reference_text", "asr_text"]
    samp_cols = [c for c in samp_cols if c in samp.columns]
    _write_df(ws6, samp[samp_cols], title="30 Random Samples – Thesis Appendix", freeze="A4")

    wb.save(EXCEL)
    logger.info("Excel saved: %s", EXCEL)


# ─────────────────────────────────────────────────────────────────────────────
# Figures
# ─────────────────────────────────────────────────────────────────────────────

def build_figures(df: pd.DataFrame):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import seaborn as sns

    sns.set_theme(style="whitegrid", font_scale=1.15)
    FIG_DIR.mkdir(parents=True, exist_ok=True)

    models = sorted(df["model"].unique())
    palette = sns.color_palette("Blues_d", len(models))

    # 1. WER boxplot
    fig, ax = plt.subplots(figsize=(9, 5))
    order = df.groupby("model")["wer"].median().sort_values().index.tolist()
    sns.boxplot(data=df, x="model", y="wer", order=order,
                palette="Blues_d", width=0.5, ax=ax, fliersize=2)
    ax.set_title("Word Error Rate (WER) – All ASR Models", fontweight="bold")
    ax.set_xlabel("Model")
    ax.set_ylabel("WER")
    ax.tick_params(axis="x", rotation=15)
    plt.tight_layout()
    fig.savefig(FIG_DIR / "wer_boxplot.png", dpi=150)
    plt.close(fig)

    # 2. WER + CER grouped bar
    fig, ax = plt.subplots(figsize=(9, 5))
    agg = df.groupby("model").agg(wer=("wer","mean")).reset_index()
    if "cer" in df.columns:
        agg["cer"] = df.groupby("model")["cer"].mean().values
        x = np.arange(len(agg))
        ax.bar(x - 0.2, agg["wer"], 0.35, label="WER", color="#2E75B6")
        ax.bar(x + 0.2, agg["cer"], 0.35, label="CER", color="#ED7D31")
        ax.set_xticks(x)
        ax.set_xticklabels(agg["model"], rotation=15)
        ax.legend()
    else:
        ax.bar(agg["model"], agg["wer"], color="#2E75B6")
        ax.set_xlabel("Model")
    ax.set_title("Mean WER and CER per ASR Model", fontweight="bold")
    ax.set_ylabel("Error Rate")
    plt.tight_layout()
    fig.savefig(FIG_DIR / "wer_cer_comparison.png", dpi=150)
    plt.close(fig)

    # 3. WER by audio duration
    if "audio_duration" in df.columns:
        fig, ax = plt.subplots(figsize=(9, 5))
        dur_bins   = [0, 5, 10, 20, 40, 999]
        dur_labels = ["0-5s", "5-10s", "10-20s", "20-40s", ">40s"]
        df2 = df.copy()
        df2["dur_bin"] = pd.cut(df2["audio_duration"], bins=dur_bins, labels=dur_labels)
        dur_agg = df2.groupby(["model","dur_bin"], observed=True)["wer"].mean().reset_index()
        for m, color in zip(models, palette):
            sub = dur_agg[dur_agg["model"] == m]
            ax.plot(sub["dur_bin"].astype(str), sub["wer"], marker="o", label=m, color=color)
        ax.set_title("WER vs Audio Duration", fontweight="bold")
        ax.set_xlabel("Audio Duration Bucket")
        ax.set_ylabel("Mean WER")
        ax.legend()
        plt.tight_layout()
        fig.savefig(FIG_DIR / "wer_vs_duration.png", dpi=150)
        plt.close(fig)

    # 4. RTF comparison
    if "processing_time" in df.columns and "audio_duration" in df.columns:
        df3 = df.copy()
        df3["rtf"] = df3["processing_time"] / df3["audio_duration"].clip(lower=0.1)
        rtf_agg = df3.groupby("model")["rtf"].mean().sort_values()
        fig, ax = plt.subplots(figsize=(8, 4))
        bars = ax.barh(rtf_agg.index, rtf_agg.values, color="#2E75B6", edgecolor="white")
        ax.bar_label(bars, fmt="%.3f", padding=3)
        ax.axvline(1.0, color="red", ls="--", lw=1.2, label="RTF = 1 (real-time)")
        ax.set_title("Real-Time Factor (RTF) per Model", fontweight="bold")
        ax.set_xlabel("RTF (lower = faster than real-time)")
        ax.legend()
        plt.tight_layout()
        fig.savefig(FIG_DIR / "rtf_comparison.png", dpi=150)
        plt.close(fig)

    # 5. Confidence histogram
    if "confidence" in df.columns:
        fig, ax = plt.subplots(figsize=(9, 5))
        for m, color in zip(models, palette):
            sub = df[df["model"] == m]["confidence"].dropna()
            ax.hist(sub, bins=40, alpha=0.55, label=m, density=True, color=color)
        ax.set_title("Confidence Score Distribution", fontweight="bold")
        ax.set_xlabel("Confidence")
        ax.set_ylabel("Density")
        ax.legend()
        plt.tight_layout()
        fig.savefig(FIG_DIR / "confidence_hist.png", dpi=150)
        plt.close(fig)

    # 6. Medical WER
    if "medical_wer" in df.columns:
        fig, ax = plt.subplots(figsize=(8, 4))
        med = df.groupby("model").agg(
            wer=("wer","mean"), medical_wer=("medical_wer","mean")
        ).reset_index().sort_values("wer")
        x = np.arange(len(med))
        ax.bar(x - 0.2, med["wer"],         0.35, label="Overall WER",  color="#2E75B6")
        ax.bar(x + 0.2, med["medical_wer"],  0.35, label="Medical WER",  color="#C00000")
        ax.set_xticks(x)
        ax.set_xticklabels(med["model"], rotation=15)
        ax.set_title("Overall WER vs Medical Term WER", fontweight="bold")
        ax.set_ylabel("Error Rate")
        ax.legend()
        plt.tight_layout()
        fig.savefig(FIG_DIR / "medical_wer_comparison.png", dpi=150)
        plt.close(fig)

    # 7. WER scatter (per sample)
    fig, axes = plt.subplots(1, len(models), figsize=(4*len(models), 4), sharey=True)
    if len(models) == 1:
        axes = [axes]
    for ax, m, color in zip(axes, models, palette):
        sub = df[df["model"] == m]
        ax.scatter(sub["audio_duration"] if "audio_duration" in sub else range(len(sub)),
                   sub["wer"], alpha=0.3, s=8, color=color)
        ax.set_title(m, fontsize=10)
        ax.set_xlabel("Duration (s)")
    axes[0].set_ylabel("WER")
    fig.suptitle("WER vs Audio Duration (per sample)", fontweight="bold")
    plt.tight_layout()
    fig.savefig(FIG_DIR / "wer_scatter.png", dpi=150)
    plt.close(fig)

    logger.info("Figures saved to %s", FIG_DIR)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    FIG_DIR.mkdir(parents=True, exist_ok=True)

    logger.info("Loading ASR results …")
    df = load_all()

    logger.info("Computing CER …")
    df = _add_cer(df)

    # Print quick summary
    summary = df.groupby("model").agg(
        samples=("wer","count"),
        mean_wer=("wer","mean"),
        median_wer=("wer","median"),
        mean_conf=("confidence","mean") if "confidence" in df.columns else ("wer","count"),
    ).round(4)
    print("\n" + "="*55)
    print("ASR EVALUATION SUMMARY")
    print("="*55)
    print(summary.sort_values("mean_wer").to_string())
    print("="*55 + "\n")

    logger.info("Building Excel workbook …")
    build_excel(df)

    logger.info("Building figures …")
    build_figures(df)

    print(f"\nResults:")
    print(f"  Excel  →  {EXCEL}")
    print(f"  Figures→  {FIG_DIR}")


if __name__ == "__main__":
    main()
