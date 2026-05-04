#!/usr/bin/env python3
"""
Script 02 – Summarization Evaluation
======================================
Takes ASR transcriptions from all_transcriptions.csv, runs all 3 summarizers,
optionally passes through GPT AI-enhancement first, and saves full results.

Evaluation flow per sample
---------------------------
  asr_text  ──────────────────────────────────► summarizer ──► summary_raw
  asr_text  ──► GPT enhancement ──► asr_ai  ──► summarizer ──► summary_ai

  Both summaries are scored against reference_text (ground truth) using:
    ROUGE-1, ROUGE-2, ROUGE-L, BLEU, Medical Term Preservation, Compression Ratio

Outputs
-------
  data/outputs/summarization_evaluation/
    all_results.csv                – every row: model × summarizer × sample
    summary_table.csv              – mean metrics per (model × summarizer) combo
    summarization_full_results.xlsx – multi-sheet Excel
    best_duo.json                  – best ASR + summarizer combination
    figures/                       – publication-quality PNG figures

Usage
-----
  # Full run (all models, all 3 summarizers, test split only)
  python scripts/02_evaluate_summarization.py

  # Limit samples for quick test
  python scripts/02_evaluate_summarization.py --samples 50

  # Run on specific split
  python scripts/02_evaluate_summarization.py --split test

  # Enable GPT AI-enhancement (needs OPENAI_API_KEY in .env)
  python scripts/02_evaluate_summarization.py --ai-enhance

  # Run only one summarizer
  python scripts/02_evaluate_summarization.py --summarizers bert_extractive
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
import warnings
from pathlib import Path

# ── silence noisy third-party loggers ────────────────────────────────────────
warnings.filterwarnings("ignore")
logging.getLogger("absl").setLevel(logging.ERROR)
logging.getLogger("tensorflow").setLevel(logging.ERROR)
logging.getLogger("transformers").setLevel(logging.WARNING)
logging.getLogger("sentence_transformers").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("datasets").setLevel(logging.WARNING)
os.environ.setdefault("TF_CPP_MIN_LOG_LEVEL", "3")

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("summarization_eval")

# ── paths ─────────────────────────────────────────────────────────────────────
# Input: produced by scripts/01_evaluate_asr.py
DATA_CSV  = ROOT / "data" / "outputs" / "asr_evaluation" / "asr_full_results.csv"

OUT_DIR   = ROOT / "data" / "outputs" / "summarization_evaluation"
CKPT_DIR  = OUT_DIR / "checkpoints"
FIG_DIR   = OUT_DIR / "figures"

# ── summarizer configs ────────────────────────────────────────────────────────
SUMMARIZERS = {
    "bert_extractive": {
        "name": "bert_extractive",
        "type": "extractive",
        "num_sentences": 3,
        "language": "de",
    },
    "bart_german": {
        "name": "bart_german",
        "type": "bart",
        "model_name": "philschmid/bart-large-german-samsum",
        "language": "de",
        "max_length": 150,
        "min_length": 25,
        "num_beams": 4,
    },
    "mt5_base": {
        "name": "mt5_base",
        "type": "mt5",
        "model_name": "google/mt5-base",
        "language": "de",
        "max_length": 128,
        "min_length": 20,
        "num_beams": 4,
    },
}


# ─────────────────────────────────────────────────────────────────────────────
# Data loading
# ─────────────────────────────────────────────────────────────────────────────

def load_transcriptions(split: str, samples: int | None) -> pd.DataFrame:
    if not DATA_CSV.exists():
        raise FileNotFoundError(
            f"ASR results not found at:\n  {DATA_CSV}\n"
            f"Run script 01 first:\n  python scripts/01_evaluate_asr.py"
        )

    df = pd.read_csv(DATA_CSV)

    # Script 01 uses 'reference' / 'hypothesis' / 'duration_s' — remap to what
    # the rest of this script expects.
    df = df.rename(columns={
        "reference":  "reference_text",
        "hypothesis": "asr_text",
        "duration_s": "audio_duration",
    })

    if split != "all":
        df = df[df["split"] == split]

    df = df.dropna(subset=["asr_text", "reference_text"])
    df = df[df["asr_text"].str.strip().str.len() > 5]
    df = df.reset_index(drop=True)

    logger.info("Loaded %d rows from asr_full_results.csv", len(df))

    if samples:
        parts = []
        for model, grp in df.groupby("asr_model"):
            parts.append(grp.head(samples))
        df = pd.concat(parts, ignore_index=True)

    logger.info("Working dataset: %d rows across models: %s",
                len(df), df["asr_model"].unique().tolist())
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Summarizer factory
# ─────────────────────────────────────────────────────────────────────────────

def build_summarizer(cfg: dict):
    from src.summarization.medical_summarizer import (
        ExtractiveSummarizer, BartSummarizer, MT5Summarizer,
    )
    t = cfg["type"]
    if t == "extractive": return ExtractiveSummarizer(cfg)
    if t == "bart":       return BartSummarizer(cfg)
    if t == "mt5":        return MT5Summarizer(cfg)
    raise ValueError(f"Unknown summarizer type: {t}")


# ─────────────────────────────────────────────────────────────────────────────
# Metrics
# ─────────────────────────────────────────────────────────────────────────────

_rouge_scorer = None

def rouge(reference: str, hypothesis: str) -> dict:
    global _rouge_scorer
    if _rouge_scorer is None:
        from rouge_score import rouge_scorer as rs
        _rouge_scorer = rs.RougeScorer(["rouge1", "rouge2", "rougeL"], use_stemmer=False)
    if not hypothesis.strip():
        return {"rouge1": 0.0, "rouge2": 0.0, "rougeL": 0.0}
    try:
        s = _rouge_scorer.score(reference, hypothesis)
        return {k: round(s[k].fmeasure, 4) for k in ("rouge1", "rouge2", "rougeL")}
    except Exception:
        return {"rouge1": 0.0, "rouge2": 0.0, "rougeL": 0.0}


def bleu(reference: str, hypothesis: str) -> float:
    try:
        import nltk
        ref = reference.lower().split()
        hyp = hypothesis.lower().split()
        if not hyp or not ref:
            return 0.0
        sf = nltk.translate.bleu_score.SmoothingFunction().method1
        return round(nltk.translate.bleu_score.sentence_bleu([ref], hyp,
               weights=(0.25,0.25,0.25,0.25), smoothing_function=sf), 4)
    except Exception:
        return 0.0


def medical_pres(source: str, summary: str) -> float:
    try:
        from src.medical.terminology import medical_term_preservation
        return round(medical_term_preservation(source, summary), 4)
    except Exception:
        return 0.0


def compression(source: str, summary: str) -> float:
    sw = len(source.split())
    return round(len(summary.split()) / sw, 4) if sw > 0 else 0.0


def score_summary(reference: str, source: str, summary: str) -> dict:
    """Score a summary against reference text (ground truth)."""
    r = rouge(reference, summary)
    return {
        "rouge1":      r["rouge1"],
        "rouge2":      r["rouge2"],
        "rougeL":      r["rougeL"],
        "bleu":        bleu(reference, summary),
        "med_pres":    medical_pres(source, summary),  # % of source medical terms kept
        "compression": compression(source, summary),
        "sum_words":   len(summary.split()),
    }


# ─────────────────────────────────────────────────────────────────────────────
# AI enhancement
# ─────────────────────────────────────────────────────────────────────────────

def build_enhancer():
    try:
        from dotenv import load_dotenv
        load_dotenv(ROOT / ".env")
    except ImportError:
        pass
    key = os.environ.get("OPENAI_API_KEY", "")
    if not key or key.startswith("your_"):
        logger.info("No OPENAI_API_KEY – AI enhancement disabled")
        return None
    try:
        from src.pipeline.ai_enhancer import AIEnhancer
        return AIEnhancer({"model": "gpt-4o-mini"})
    except Exception as exc:
        logger.warning("AIEnhancer load failed: %s", exc)
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Checkpoint
# ─────────────────────────────────────────────────────────────────────────────

def ckpt_file(asr_model: str, summ_name: str) -> Path:
    return CKPT_DIR / f"{asr_model}__{summ_name}.csv"


def load_ckpt(asr_model: str, summ_name: str) -> tuple[list, set]:
    p = ckpt_file(asr_model, summ_name)
    if p.exists():
        try:
            df = pd.read_csv(p)
            logger.info("  Checkpoint: %d rows already done", len(df))
            return df.to_dict("records"), set(df["sample_id"].tolist())
        except Exception:
            pass
    return [], set()


def save_ckpt(rows: list, asr_model: str, summ_name: str):
    if rows:
        pd.DataFrame(rows).to_csv(ckpt_file(asr_model, summ_name), index=False)


# ─────────────────────────────────────────────────────────────────────────────
# Core loop
# ─────────────────────────────────────────────────────────────────────────────

def run_combo(df: pd.DataFrame, asr_model: str, summ_name: str,
              summ_cfg: dict, enhancer, split_name: str) -> pd.DataFrame:

    rows, done_ids = load_ckpt(asr_model, summ_name)
    subset = df[df["asr_model"] == asr_model].copy()
    todo   = subset[~subset["sample_id"].isin(done_ids)]

    logger.info("[%s × %s]  %d total  /  %d done  /  %d to run",
                asr_model, summ_name, len(subset), len(done_ids), len(todo))

    if todo.empty:
        return pd.DataFrame(rows)

    try:
        summ = build_summarizer(summ_cfg)
        summ.load_model()
    except Exception as exc:
        logger.error("  Failed to load %s: %s", summ_name, exc)
        return pd.DataFrame(rows)

    for count, (_, row) in enumerate(todo.iterrows(), 1):
        asr_text = str(row.get("asr_text", "")).strip()
        ref_text = str(row.get("reference_text", "")).strip()
        sid      = int(row.get("sample_id", 0))
        wer      = float(row.get("wer", 0.0)) if pd.notna(row.get("wer")) else None
        conf     = float(row.get("confidence", 0.0)) if pd.notna(row.get("confidence")) else None
        dur      = float(row.get("audio_duration", 0.0)) if pd.notna(row.get("audio_duration")) else None

        rec = {
            "asr_model":   asr_model,
            "summarizer":  summ_name,
            "split":       split_name,
            "sample_id":   sid,
            "asr_wer":     round(wer, 4) if wer is not None else None,
            "confidence":  round(conf, 4) if conf is not None else None,
            "audio_dur_s": round(dur, 2) if dur is not None else None,
            "asr_words":   len(asr_text.split()),
            "ref_words":   len(ref_text.split()),
            # full texts for manual inspection
            "reference_text": ref_text,
            "asr_text":       asr_text,
        }

        # ── summarize raw ASR ─────────────────────────────────────────────
        t0 = time.time()
        try:
            res = summ.summarize(asr_text)
            summary_raw = res.summary.strip()
        except Exception as exc:
            logger.debug("  [%d] summarize failed: %s", sid, exc)
            summary_raw = ""
        proc_time = round(time.time() - t0, 3)

        s_raw = score_summary(ref_text, asr_text, summary_raw)
        rec.update({
            "summary_raw":        summary_raw,
            "proc_time_s":        proc_time,
            "rouge1_raw":         s_raw["rouge1"],
            "rouge2_raw":         s_raw["rouge2"],
            "rougeL_raw":         s_raw["rougeL"],
            "bleu_raw":           s_raw["bleu"],
            "med_pres_raw":       s_raw["med_pres"],
            "compression_raw":    s_raw["compression"],
            "sum_words_raw":      s_raw["sum_words"],
        })

        # ── AI-enhance then summarize ─────────────────────────────────────
        if enhancer is not None:
            try:
                enh = enhancer.enhance(asr_text)
                ai_text = enh.enhanced_text.strip() if enh.enhanced_text else asr_text
            except Exception:
                ai_text = asr_text

            try:
                res_ai = summ.summarize(ai_text)
                summary_ai = res_ai.summary.strip()
            except Exception:
                summary_ai = ""

            s_ai = score_summary(ref_text, ai_text, summary_ai)
            rec.update({
                "asr_text_ai":     ai_text,
                "summary_ai":      summary_ai,
                "rouge1_ai":       s_ai["rouge1"],
                "rouge2_ai":       s_ai["rouge2"],
                "rougeL_ai":       s_ai["rougeL"],
                "bleu_ai":         s_ai["bleu"],
                "med_pres_ai":     s_ai["med_pres"],
                "compression_ai":  s_ai["compression"],
                "sum_words_ai":    s_ai["sum_words"],
                "delta_rougeL":    round(s_ai["rougeL"] - s_raw["rougeL"], 4),
                "delta_bleu":      round(s_ai["bleu"] - s_raw["bleu"], 4),
                "delta_med_pres":  round(s_ai["med_pres"] - s_raw["med_pres"], 4),
            })

        rows.append(rec)

        if count % 25 == 0:
            save_ckpt(rows, asr_model, summ_name)
            logger.info("  → checkpoint: %d / %d", len(rows), len(subset))

    save_ckpt(rows, asr_model, summ_name)

    # free model memory
    try:
        for attr in ("_model", "_tokenizer", "_sentence_encoder"):
            if hasattr(summ, attr):
                setattr(summ, attr, None)
        import gc, torch
        gc.collect()
        if hasattr(torch.backends, "mps") and torch.backends.mps.is_available():
            torch.mps.empty_cache()
        elif torch.cuda.is_available():
            torch.cuda.empty_cache()
    except Exception:
        pass

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────────────────────

def _autofit(ws):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 60)


def _write(ws, df, title=None, freeze="A2"):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    HF = PatternFill("solid", fgColor="1F4E79")
    HT = Font(color="FFFFFF", bold=True, size=10)
    TF = Font(bold=True, size=13)
    start = 1
    if title:
        ws.cell(1, 1, title).font = TF
        start = 3
    for ri, row in enumerate(dataframe_to_rows(df, index=False, header=True), start):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            if ri == start:
                cell.font = HT; cell.fill = HF
                cell.alignment = Alignment(horizontal="center")
    if freeze:
        ws.freeze_panes = freeze
    _autofit(ws)


def build_excel(all_df: pd.DataFrame, has_ai: bool, out_path: Path):
    import openpyxl
    wb = openpyxl.Workbook()

    metric_raw = ["rouge1_raw", "rouge2_raw", "rougeL_raw", "bleu_raw",
                  "med_pres_raw", "compression_raw", "proc_time_s"]

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Summary"
    agg = {}
    for c in metric_raw:
        if c in all_df.columns:
            agg[f"{c}_mean"] = (c, "mean")
            agg[f"{c}_std"]  = (c, "std")
    if has_ai:
        for c in ["rougeL_ai", "bleu_ai", "med_pres_ai",
                  "delta_rougeL", "delta_bleu", "delta_med_pres"]:
            if c in all_df.columns:
                agg[f"{c}_mean"] = (c, "mean")

    summ_df = all_df.groupby(["asr_model", "summarizer"]).agg(**agg).reset_index()

    # composite score: 0.45*rougeL + 0.30*med_pres + 0.15*bleu + 0.10*(1-wer)
    if "rougeL_raw_mean" in summ_df.columns:
        wer_map = all_df.groupby("asr_model")["asr_wer"].mean()
        summ_df["mean_asr_wer"] = summ_df["asr_model"].map(wer_map).round(4)
        summ_df["composite"] = (
            0.45 * summ_df["rougeL_raw_mean"].fillna(0)
          + 0.30 * summ_df["med_pres_raw_mean"].fillna(0)
          + 0.15 * summ_df["bleu_raw_mean"].fillna(0)
          + 0.10 * (1 - summ_df["mean_asr_wer"].fillna(1))
        ).round(4)
        summ_df = summ_df.sort_values("composite", ascending=False)

    float_c = [c for c in summ_df.columns if summ_df[c].dtype == float]
    summ_df[float_c] = summ_df[float_c].round(4)
    summ_df.insert(0, "rank", range(1, len(summ_df) + 1))
    _write(ws, summ_df, "Summarization Evaluation – Ranked Combinations", freeze="A4")

    # ── Sheet 2: Per-Sample Results ───────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Sample")
    show = ["asr_model", "summarizer", "split", "sample_id", "asr_wer",
            "asr_words", "ref_words", "sum_words_raw",
            "rouge1_raw", "rouge2_raw", "rougeL_raw", "bleu_raw",
            "med_pres_raw", "compression_raw", "proc_time_s"]
    if has_ai:
        show += ["rougeL_ai", "bleu_ai", "med_pres_ai",
                 "delta_rougeL", "delta_bleu", "delta_med_pres"]
    show = [c for c in show if c in all_df.columns]
    per = all_df[show].sort_values(["asr_model", "summarizer", "sample_id"])
    float_c = [c for c in per.columns if per[c].dtype == float]
    per[float_c] = per[float_c].round(4)
    _write(ws2, per, freeze="A2")

    # ── Sheet 3: AI Impact ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("AI-Impact")
    if has_ai:
        delta_cols = [c for c in ["delta_rougeL","delta_bleu","delta_med_pres"] if c in all_df.columns]
        if delta_cols:
            ai_agg = {f"{c}_mean": (c, "mean") for c in delta_cols}
            ai_agg.update({f"{c}_positive_pct": (c, lambda x: (x > 0).mean() * 100) for c in delta_cols})
            ai_df = all_df.groupby(["asr_model","summarizer"]).agg(**ai_agg).reset_index()
            ai_df = ai_df.round(4)
            _write(ws3, ai_df, "AI Enhancement Impact (positive = AI helped)", freeze="A2")
    else:
        ws3.cell(1, 1, "AI enhancement was not enabled in this run (no OPENAI_API_KEY).")

    # ── Sheet 4: Best Combinations ───────────────────────────────────────────
    ws4 = wb.create_sheet("Best-Combinations")
    best_cols = ["rank", "asr_model", "summarizer", "composite",
                 "rougeL_raw_mean", "bleu_raw_mean", "med_pres_raw_mean", "mean_asr_wer"]
    best_cols = [c for c in best_cols if c in summ_df.columns]
    _write(ws4, summ_df[best_cols].head(10), "Top 10 Combinations (Composite Score)", freeze="A4")

    # ── Sheet 5: Full Text Samples (for manual inspection) ───────────────────
    ws5 = wb.create_sheet("Manual-Inspection")
    # best combo only, sorted by rougeL descending
    if "composite" in summ_df.columns and len(summ_df) > 0:
        best_model = summ_df.iloc[0]["asr_model"]
        best_summ  = summ_df.iloc[0]["summarizer"]
        sample_df  = all_df[
            (all_df["asr_model"] == best_model) &
            (all_df["summarizer"] == best_summ)
        ].nlargest(50, "rougeL_raw")
    else:
        sample_df = all_df.head(50)

    text_cols = ["sample_id", "asr_wer", "rougeL_raw", "med_pres_raw",
                 "reference_text", "asr_text", "summary_raw"]
    if has_ai and "summary_ai" in all_df.columns:
        text_cols.append("summary_ai")
    text_cols = [c for c in text_cols if c in sample_df.columns]
    _write(ws5, sample_df[text_cols], "Top 50 Samples – Best Combo (for Manual Review)", freeze="A4")

    # ── Sheet 6: Worst Samples (for error analysis) ───────────────────────────
    ws6 = wb.create_sheet("Error-Analysis")
    worst = all_df.nsmallest(100, "rougeL_raw")[
        [c for c in ["asr_model","summarizer","sample_id","asr_wer",
                     "rougeL_raw","med_pres_raw","reference_text","asr_text","summary_raw"]
         if c in all_df.columns]
    ]
    _write(ws6, worst, "100 Worst Samples by ROUGE-L", freeze="A4")

    wb.save(out_path)
    logger.info("Excel saved: %s", out_path)
    return summ_df


# ─────────────────────────────────────────────────────────────────────────────
# Figures
# ─────────────────────────────────────────────────────────────────────────────

def build_figures(all_df: pd.DataFrame, summ_df: pd.DataFrame, has_ai: bool):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import seaborn as sns

    logging.getLogger("matplotlib").setLevel(logging.WARNING)
    sns.set_theme(style="whitegrid", font_scale=1.1)
    FIG_DIR.mkdir(parents=True, exist_ok=True)

    # 1. ROUGE-L heatmap
    fig, ax = plt.subplots(figsize=(9, 5))
    piv = all_df.groupby(["asr_model","summarizer"])["rougeL_raw"].mean().unstack(fill_value=0)
    sns.heatmap(piv, annot=True, fmt=".3f", cmap="YlOrRd", linewidths=0.5, ax=ax,
                cbar_kws={"label":"ROUGE-L (mean)"})
    ax.set_title("ROUGE-L: ASR Model × Summarizer", fontweight="bold", pad=12)
    plt.tight_layout(); fig.savefig(FIG_DIR/"rouge_heatmap.png", dpi=150); plt.close(fig)

    # 2. Medical term preservation
    fig, ax = plt.subplots(figsize=(11, 5))
    mp = all_df.groupby(["asr_model","summarizer"])["med_pres_raw"].mean().reset_index()
    mp["combo"] = mp["asr_model"] + "\n" + mp["summarizer"]
    bars = ax.bar(mp["combo"], mp["med_pres_raw"],
                  color=sns.color_palette("Blues_d", len(mp)), edgecolor="white")
    ax.bar_label(bars, fmt="%.3f", padding=3, fontsize=9)
    ax.set_ylim(0, 1.15)
    ax.set_title("Medical Term Preservation per Combination", fontweight="bold")
    ax.set_ylabel("Medical Term Preservation (0–1)")
    ax.tick_params(axis="x", rotation=30)
    plt.tight_layout(); fig.savefig(FIG_DIR/"medical_preservation.png", dpi=150); plt.close(fig)

    # 3. AI impact delta (if available)
    if has_ai and "rougeL_ai" in all_df.columns:
        fig, ax = plt.subplots(figsize=(11, 5))
        cmp = all_df.groupby(["asr_model","summarizer"]).agg(
            raw=("rougeL_raw","mean"), ai=("rougeL_ai","mean")
        ).reset_index().dropna(subset=["ai"])
        cmp["combo"] = cmp["asr_model"] + "\n" + cmp["summarizer"]
        x = np.arange(len(cmp)); w = 0.35
        ax.bar(x-w/2, cmp["raw"], w, label="Raw ASR",     color="#4472C4")
        ax.bar(x+w/2, cmp["ai"],  w, label="AI-Enhanced", color="#ED7D31")
        ax.set_xticks(x); ax.set_xticklabels(cmp["combo"], rotation=30, ha="right")
        ax.set_ylabel("Mean ROUGE-L")
        ax.set_title("ROUGE-L: Raw vs AI-Enhanced Transcription", fontweight="bold")
        ax.legend(); plt.tight_layout()
        fig.savefig(FIG_DIR/"ai_impact.png", dpi=150); plt.close(fig)

    # 4. WER vs ROUGE-L scatter
    fig, ax = plt.subplots(figsize=(8, 6))
    for name, grp in all_df.groupby("summarizer"):
        s = grp.sample(min(400, len(grp)), random_state=42)
        ax.scatter(s["asr_wer"], s["rougeL_raw"], alpha=0.35, s=15, label=name)
    ax.set_xlabel("ASR Word Error Rate"); ax.set_ylabel("Summarizer ROUGE-L")
    ax.set_title("ASR Quality vs Summarization Quality", fontweight="bold")
    ax.legend(title="Summarizer")
    try:
        sub = all_df[["asr_wer","rougeL_raw"]].dropna().sample(min(2000,len(all_df)),random_state=0)
        m, b = np.polyfit(sub["asr_wer"], sub["rougeL_raw"], 1)
        xs = np.linspace(sub["asr_wer"].min(), sub["asr_wer"].max(), 100)
        ax.plot(xs, m*xs+b, "k--", lw=1.5, label=f"trend (slope={m:.3f})")
        ax.legend(title="Summarizer")
    except Exception:
        pass
    plt.tight_layout(); fig.savefig(FIG_DIR/"wer_vs_rouge.png", dpi=150); plt.close(fig)

    # 5. BLEU comparison
    fig, ax = plt.subplots(figsize=(10, 5))
    bl = all_df.groupby(["asr_model","summarizer"])["bleu_raw"].mean().unstack(fill_value=0)
    bl.plot(kind="bar", ax=ax, colormap="tab10", edgecolor="white")
    ax.set_title("BLEU Score per Combination", fontweight="bold")
    ax.set_ylabel("Mean BLEU")
    ax.tick_params(axis="x", rotation=30)
    ax.legend(title="Summarizer", bbox_to_anchor=(1.01,1), loc="upper left")
    plt.tight_layout(); fig.savefig(FIG_DIR/"bleu_comparison.png", dpi=150); plt.close(fig)

    # 6. Compression ratio boxplot
    fig, ax = plt.subplots(figsize=(9, 5))
    all_df.boxplot(column="compression_raw", by="summarizer", ax=ax,
                   patch_artist=True,
                   boxprops=dict(facecolor="#4472C4", color="#1F4E79"),
                   medianprops=dict(color="orange", linewidth=2),
                   flierprops=dict(marker=".", markersize=3))
    ax.set_title("Compression Ratio by Summarizer", fontweight="bold")
    plt.suptitle(""); ax.set_xlabel("Summarizer"); ax.set_ylabel("Compression Ratio")
    plt.tight_layout(); fig.savefig(FIG_DIR/"compression_boxplot.png", dpi=150); plt.close(fig)

    # 7. Summary length histogram
    fig, ax = plt.subplots(figsize=(9, 5))
    for name, grp in all_df.groupby("summarizer"):
        ax.hist(grp["sum_words_raw"], bins=40, alpha=0.55, label=name, density=True)
    ax.set_xlabel("Summary Length (words)"); ax.set_ylabel("Density")
    ax.set_title("Distribution of Summary Lengths", fontweight="bold")
    ax.legend(title="Summarizer")
    plt.tight_layout(); fig.savefig(FIG_DIR/"summary_length_hist.png", dpi=150); plt.close(fig)

    logger.info("Figures saved: %s", FIG_DIR)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--split",       default="test",
                   choices=["train","eval","test","all"],
                   help="Which dataset split to use (default: test)")
    p.add_argument("--samples",     type=int, default=None,
                   help="Limit samples per ASR model (default: all)")
    p.add_argument("--summarizers", nargs="+",
                   default=list(SUMMARIZERS.keys()),
                   choices=list(SUMMARIZERS.keys()),
                   help="Which summarizers to run")
    p.add_argument("--ai-enhance",  action="store_true",
                   help="Enable GPT AI-enhancement step (needs OPENAI_API_KEY)")
    p.add_argument("--no-figures",  action="store_true")
    return p.parse_args()


def main():
    args = parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    CKPT_DIR.mkdir(parents=True, exist_ok=True)

    try:
        import nltk
        for pkg in ("punkt", "punkt_tab"):
            nltk.download(pkg, quiet=True)
    except Exception:
        pass

    logger.info("=" * 55)
    logger.info("Script 02 – Summarization Evaluation")
    logger.info("Split      : %s", args.split)
    logger.info("Samples    : %s", args.samples or "ALL")
    logger.info("Summarizers: %s", args.summarizers)
    logger.info("AI Enhance : %s", args.ai_enhance)
    logger.info("=" * 55)

    # Load data
    data = load_transcriptions(args.split, args.samples)
    asr_models = data["asr_model"].unique().tolist()

    # AI enhancer
    enhancer = build_enhancer() if args.ai_enhance else None
    has_ai   = enhancer is not None

    # Run all combinations
    all_frames = []
    for asr_model in asr_models:
        for summ_name in args.summarizers:
            logger.info("")
            t0 = time.time()
            result = run_combo(data, asr_model, summ_name,
                               SUMMARIZERS[summ_name], enhancer, args.split)
            if result is not None and not result.empty:
                all_frames.append(result)
                elapsed = time.time() - t0
                logger.info("  Done: %d rows | ROUGE-L=%.3f | MedPres=%.3f | %.0fs",
                            len(result),
                            result["rougeL_raw"].mean() if "rougeL_raw" in result else 0,
                            result["med_pres_raw"].mean() if "med_pres_raw" in result else 0,
                            elapsed)

    if not all_frames:
        logger.error("No results produced.")
        return 1

    all_df = pd.concat(all_frames, ignore_index=True)

    # Save raw CSV (full text included for manual checking)
    csv_out = OUT_DIR / "all_results.csv"
    all_df.to_csv(csv_out, index=False)
    logger.info("Raw CSV: %s", csv_out)

    # Build Excel
    logger.info("Building Excel …")
    excel_out = OUT_DIR / "summarization_full_results.xlsx"
    summ_df = build_excel(all_df, has_ai, excel_out)

    # Print ranking
    print("\n" + "=" * 60)
    print("SUMMARIZATION EVALUATION – FINAL RANKING")
    print("=" * 60)
    show_cols = ["rank","asr_model","summarizer","composite",
                 "rougeL_raw_mean","bleu_raw_mean","med_pres_raw_mean","mean_asr_wer"]
    show_cols = [c for c in show_cols if c in summ_df.columns]
    print(summ_df[show_cols].to_string(index=False))
    print("=" * 60)

    # Save best_duo.json
    if len(summ_df) > 0 and "composite" in summ_df.columns:
        best = summ_df.iloc[0]
        best_json = OUT_DIR / "best_duo.json"
        best_json.write_text(json.dumps({
            "best_asr_model":  str(best["asr_model"]),
            "best_summarizer": str(best["summarizer"]),
            "composite_score": float(best.get("composite", 0)),
            "rougeL":          float(best.get("rougeL_raw_mean", 0)),
            "bleu":            float(best.get("bleu_raw_mean", 0)),
            "medical_pres":    float(best.get("med_pres_raw_mean", 0)),
            "mean_asr_wer":    float(best.get("mean_asr_wer", 0)),
            "ranking": [
                {"rank": int(r["rank"]), "asr_model": str(r["asr_model"]),
                 "summarizer": str(r["summarizer"]),
                 "composite": float(r.get("composite",0))}
                for _, r in summ_df.head(10).iterrows()
            ]
        }, indent=2, ensure_ascii=False))
        logger.info("Best duo: %s + %s (score=%.4f)",
                    best["asr_model"], best["summarizer"], best.get("composite", 0))

    # Build figures
    if not args.no_figures:
        logger.info("Building figures …")
        build_figures(all_df, summ_df, has_ai)

    print(f"\n Results saved to:")
    print(f"   CSV   →  {csv_out}")
    print(f"   Excel →  {excel_out}")
    print(f"   Figs  →  {FIG_DIR}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
